import openai
from openpyxl import Workbook
from openpyxl import load_workbook
import time

#all of this software is suitable to be used with the version of Anki Versión 2.1.65 (aa9a734f)

# Genera una API Key desde https://openai.com/api
openai.api_key = "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"


def ChatGPT_Word(pregunta):

    prompt = pregunta

    completion = openai.Completion.create(engine="text-davinci-003",
                                          prompt=prompt,
                                          max_tokens=512)
    
    print("ChatGPT_Word function was used")

    return completion.choices[0].text

def ChatGPT_Sentence(pregunta):

    prompt = pregunta

    completion = openai.Completion.create(engine="text-davinci-003",
                                          prompt=prompt,
                                          max_tokens=2048)
    
    print("ChatGPT_Sentence function was used")

    return completion.choices[0].text




document = 'Libro1.xlsx'

Book = load_workbook(document)

Page = Book['Aleman__Sustantivos']

C_Sustantivo_Español = 4            #1
C_Sustantivo_Aleman = 5             #2
C_Sustantivo_Aleman_Plural = 6      #3
C_Frase_Aleman = 7                  #4
C_Frase_Español = 8                 #5




contadorVertical = 1

while (str(Page.cell(contadorVertical,C_Sustantivo_Español).value) != 'None' or str(Page.cell(contadorVertical,C_Sustantivo_Aleman).value) != 'None'):

    #fill the column of "Sustantivo aleman"
    if (str(Page.cell(contadorVertical,C_Sustantivo_Aleman).value) == 'None'):

        Page.cell(contadorVertical,C_Sustantivo_Aleman).value = ChatGPT_Word("Traduceme esta palabra del español al aleman : " + str(Page.cell(contadorVertical,C_Sustantivo_Español).value))
        time.sleep(20)
        print("*****************************************************************************")
        print(str(Page.cell(contadorVertical,C_Sustantivo_Español).value) + " - " + str(Page.cell(contadorVertical,C_Sustantivo_Aleman).value))
        print("*****************************************************************************")

    #fill the column of "Palabra plural en aleman"
    if (str(Page.cell(contadorVertical,C_Sustantivo_Aleman_Plural).value) == 'None'):

        respuesta = ChatGPT_Word("schreibt den Plural von " + str(Page.cell(contadorVertical,C_Sustantivo_Aleman).value) + "mit seinem Artikel")
        
        if len(respuesta.split(" ")) == 2 or len(str(Page.cell(contadorVertical,C_Sustantivo_Español).value).split(" ")) != 2:
            print("*****************************************************************************")
            print("la respuesta de gpt es correta")
            Page.cell(contadorVertical,C_Sustantivo_Aleman_Plural).value = respuesta
            print(str(Page.cell(contadorVertical,C_Sustantivo_Aleman).value) + " - " + str(Page.cell(contadorVertical,C_Sustantivo_Aleman_Plural).value))
            print("*****************************************************************************")

        else:
            print("*****************************************************************************") 
            print("la respuesta de gpt no es correta, será descartada: " + respuesta)
            print("*****************************************************************************")
       
        time.sleep(20)
    
    #fill the column of "Frase en Aleman"
    if (str(Page.cell(contadorVertical,C_Frase_Aleman).value) == 'None'):
        FraseAleman = ChatGPT_Sentence("dime una frase en aleman con la palabra (" + str(Page.cell(contadorVertical,C_Sustantivo_Aleman).value) + "escribeme solo la frase, nada mas")
        Page.cell(contadorVertical,C_Frase_Aleman).value = FraseAleman
        print("*****************************************************************************")
        print(FraseAleman)
        print("*****************************************************************************")
        time.sleep(20)

    #fill the column of frase en español
    if (str(Page.cell(contadorVertical,C_Frase_Español).value) == 'None'):
        FraseEspañol = ChatGPT_Sentence("Traduce la frase al español: " + str(Page.cell(contadorVertical,C_Frase_Aleman).value) + "escribeme solo la frase, nada mas")
        Page.cell(contadorVertical,C_Frase_Español).value = FraseEspañol
        print("*****************************************************************************")
        print(FraseEspañol)
        print("*****************************************************************************")
        time.sleep(20)

    #format corrector, the format should be this 'das Haus'
#
#
#
#
#
#
#

    contadorVertical += 1
    Book.save('Libro1.xlsx')


Book.save('Libro1.xlsx')
Book.close()